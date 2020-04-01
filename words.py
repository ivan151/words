from flask import Flask, request
from random import choice
import telepot
import urllib3
import time
from openpyxl import Workbook, load_workbook
import os.path
from langdetect import detect

# import datetime

proxy_url = "http://proxy.server:3128"
telepot.api._pools = {
    'default': urllib3.ProxyManager(proxy_url=proxy_url, num_pools=3, maxsize=10, retries=False, timeout=30),
}
telepot.api._onetime_pool_spec = (
    urllib3.ProxyManager, dict(proxy_url=proxy_url, num_pools=1, maxsize=1, retries=False, timeout=30))

secret = "odBabTAsR6XGawl54vfm1P9aU8A8hfCh"
bot = telepot.Bot('1070561990:AAEsauHeiN-Yh31KeDZr9Y2gsz3SaX32Grw')
bot.setWebhook("https://words151.pythonanywhere.com/{}".format(secret), max_connections=1)


# Takes a random word from file

def random_word(filename):
    with open(filename, 'r') as words:
        random_word = choice(words.readlines())
        words.close()
        return random_word


# Create a new workbook with three column
def create_wb(chat_id):
    file_name = f'words/{chat_id}' + '.xlsx'
    wb = Workbook()
    sheet = wb['Sheet']
    sheet2 = wb.create_sheet('Sheet2')
    sheet['A1'].value = 'chat_id'
    sheet['B1'].value = 'user_name'
    sheet['C1'].value = 'datetime'
    sheet['D1'].value = 'word'
    sheet2['A1'].value = 'chat_id'
    sheet2['B1'].value = 'starttime'
    sheet2['C1'].value = 'stoptime'
    sheet2['D1'].value = 'origin_word'
    wb.save(filename=file_name)


# Insert player's word to workbook
def insert_word_to_wb(chat_id, user_name, time_info, word):
    file_name = f'words/{chat_id}' + '.xlsx'
    wb = load_workbook(file_name)
    sheet = wb['Sheet']
    sheet[sheet.max_row + 1][0].value = chat_id
    sheet[sheet.max_row][1].value = user_name
    sheet[sheet.max_row][2].value = time_info
    sheet[sheet.max_row][3].value = word
    wb.save(filename=file_name)


# Insert starttime and origin word to wb
def insert_start_time(chat_id, word):
    file_name = f'words/{chat_id}' + '.xlsx'
    start_time = time.time()
    wb = load_workbook(file_name)
    sheet = wb['Sheet2']
    sheet[sheet.max_row + 1][0].value = chat_id
    sheet[sheet.max_row][1].value = start_time
    # sheet[sheet.max_row + 1][2].value = stop_time
    sheet[sheet.max_row][3].value = word
    wb.save(filename=file_name)


def insert_stop_time(chat_id, word):
    file_name = f'words/{chat_id}' + '.xlsx'
    stop_time = time.time()
    wb = load_workbook(file_name)
    sheet = wb['Sheet2']
    sheet[sheet.max_row][2].value = stop_time
    wb.save(filename=file_name)


def get_result(chat_id):
    file_name = f'words/{chat_id}' + '.xlsx'
    wb = load_workbook(file_name)
    sheet = wb['Sheet']
    sheet2 = wb['Sheet2']
    start_time = int(sheet2[sheet2.max_row][1].value)
    stop_time = int(sheet2[sheet2.max_row][2].value)
    word_time = sheet['C']
    #cold = sheet['D']
    words = {}
    for cell in word_time:
        if cell.value in range(start_time,stop_time):
            word = sheet.cell(row=cell.row, column=4).value
            user = sheet.cell(row=cell.row, column=2).value
            if user in words:
                words[user].append([cell.value,word])
            else:
                words[user] = [[cell.value,word]]
    return words


def right_word(chat_id, word):
    file_name = f'words/{chat_id}' + '.xlsx'
    wb = load_workbook(file_name)
    sheet = wb['Sheet2']
    origin_word = sheet[sheet.max_row][3].value
    count = 0
    for letter in word.lower():
        if word.count(letter) > origin_word.count(letter):
            continue
        elif origin_word.count(letter) == 0:
            continue
        else:
            count += 1
    if count == len(word):
        return True
    else:
        return False


def whether_exists(word):
    lang = detect(word)
    if lang == 'en':
        word = word.lower() + '\n'
        with open('words/words.txt', 'r') as file:
            words = file.readlines()
            if word in words:
                return True
            else:
                return False
    elif lang == 'ru':
        word = word.lower() + '\n'
        with open('words/rus.txt', 'r') as file:
            words = file.readlines()
            if word in words:
                return True
            else:
                return False


def inter(words):
    values = list(words.values())
    a = values[0]
    b = values[1]
    da = dict((e[::-1] for e in a))
    db = dict((e[::-1] for e in b))
    intersection_list = [
            [max(da[k], db[k]), k] for k in set(da).intersection(db)
            ]
    return intersection_list


def checker(user, words):
    user_words = words[user]
    duplicates = []
    for word in user_words:
        if word in inter(words):
            duplicates.append(word)
            user_words.remove(word)
    return [user_words, duplicates]


app = Flask(__name__)


@app.route('/{}'.format(secret), methods=["POST"])
def telegram_webhook():
    update = request.get_json()
    if "message" in update:
        chat_id = update["message"]["chat"]["id"]
        # user_id = update['message']['from']['id']
        user_name = update['message']['from']['username']
        if os.path.isfile(f'words/{chat_id}' + '.xlsx'):
            if "text" in update["message"]:
                text = update["message"]["text"]
                time_info = update["message"]["date"]

                if "/starteng" in text:
                    # bot gives a word to players
                    word = random_word('words/long_words.txt')
                    insert_start_time(chat_id, word)
                    bot.sendMessage(chat_id, "{}".format(word))
                    # game begins, timer is started
                    mins = 0
                    while mins < 5:
                        bot.sendMessage(chat_id, 'Word: {} >>>>>>>> {} minutes left'.format(word,mins))
                        time.sleep(5)
                        mins += 1
                    else:
                        insert_stop_time(chat_id, word)
                        bot.sendMessage(chat_id, 'Time is over!')

                if "/startrus" in text:
                    # bot gives a word to players
                    word = random_word('words/long_words_rus.txt')
                    insert_start_time(chat_id, word)
                    bot.sendMessage(chat_id, "{}".format(word))
                    # game begins, timer is started
                    mins = 0
                    while mins < 5:
                        bot.sendMessage(chat_id, 'Word: {} >>>>>>>> {} minutes left'.format(word,mins))
                        time.sleep(5)
                        mins += 1
                    else:
                        insert_stop_time(chat_id, word)
                        bot.sendMessage(chat_id, 'Time is over!')
                elif text == '/result':
                    try:
                        words = get_result(chat_id)
                        if len(words) < 2:
                            bot.sendMessage(chat_id, '''You don't have a result, 
                                            because you don't have a partner!''')
                        elif len(words) > 2:
                            bot.sendMessage(chat_id, "This game created only for two players!")
                        else:
                            for key in words.keys():
                                user_words = checker(key, words)
                                duplicates = ''
                                for word in user_words[1]:
                                    duplicates += f'{word[1]}'
                                suitable_words = []
                                suitable_and_existing = []
                                unsuitable_words = []
                                doesnt_exist = []
                                for word in user_words[0]:
                                    if right_word(chat_id,word[1]):
                                            suitable_words.append(word[1])
                                    else:
                                        unsuitable_words.append(word)
                                for word in suitable_words:
                                    if whether_exists(word):
                                        suitable_and_existing.append(word)
                                    else:
                                        doesnt_exist.append(word)

                                points = 0
                                for word in suitable_and_existing:
                                    points += len(word)
                                bot.sendMessage(chat_id,
                                                '''Player: {player} 
                                                | Correct words: {correct}  
                                                |  Not nested words: {not_nested} 
                                                | Doesn't exist: {doesnt} 
                                                |  Duplicates: {dup} 
                                                | Points: {p}'''.format(
                                                   player = key,
                                                   correct = ' '.join(map(str, suitable_and_existing)),
                                                   not_nested = ' '.join(map(str, unsuitable_words)),
                                                   doesnt = ' '.join(map(str, doesnt_exist)),
                                                   dup = ' '.join(map(str, duplicates)),
                                                   p = points))
                    except:
                        bot.sendMessage(chat_id, "You haven't played yet!")



                else:
                    word = text
                    insert_word_to_wb(chat_id, user_name,time_info, word)

            else:
                bot.sendMessage(chat_id, "it's not text!")
        else:
            create_wb(chat_id)
            bot.sendMessage(chat_id, "You a new player, press start to run the game")

    return "OK"
